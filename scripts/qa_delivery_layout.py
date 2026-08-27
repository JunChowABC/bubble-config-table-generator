from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("需要 Python 3 和 openpyxl 才能运行此验证器") from exc


VALID_ROLES = {"feature", "shared", "referenced"}
VALID_ACTIONS = {"created", "copied_and_modified"}
VALID_TEST_OPERATIONS = {"added", "updated"}
VALID_ID_SCOPES = {"sheet", "module", "parent"}
VALID_ID_KINDS = {"new", "derived_child", "test", "reused", "updated"}
VALID_ID_SOURCES = {"S", "T", "D", "A"}
VALID_ID_STATUSES = {"reused", "candidate", "confirmed"}
INT32_MAX = 2_147_483_647


def normalized(path: Path) -> str:
    return os.path.normcase(os.path.normpath(str(path.resolve())))


def same_id(value: Any, expected: Any) -> bool:
    if value is None or expected is None:
        return False
    return str(value).strip() == str(expected).strip()


def rows_for_id(worksheet, expected_id: Any) -> list[tuple[Any, ...]]:
    matches = []
    for row in worksheet.iter_rows(min_row=7, values_only=True):
        if row and str(row[0] or "").strip().lower() == "end":
            break
        if len(row) > 1 and same_id(row[1], expected_id):
            matches.append(tuple(row))
    return matches


def main() -> int:
    parser = argparse.ArgumentParser(description="验证 Bubble 配置表的输出路径与工作簿聚合")
    parser.add_argument("manifest", type=Path, help="generation-manifest.json")
    args = parser.parse_args()

    manifest_path = args.manifest.resolve()
    if not manifest_path.is_file():
        raise SystemExit(f"交付清单不存在: {manifest_path}")

    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    def add(bucket, code: str, message: str, **details):
        bucket.append({"code": code, "message": message, **details})

    resolved_value = data.get("resolved_output_directory")
    if not resolved_value:
        add(errors, "MISSING_OUTPUT_DIRECTORY", "缺少resolved_output_directory")
        output_dir = manifest_path.parent
    else:
        output_dir = Path(resolved_value).expanduser().resolve()
        if not output_dir.is_dir():
            add(errors, "OUTPUT_DIRECTORY_NOT_FOUND", "实际输出目录不存在", path=str(output_dir))

    if normalized(manifest_path.parent) != normalized(output_dir):
        add(errors, "MANIFEST_OUTSIDE_OUTPUT_DIRECTORY", "generation-manifest.json必须位于实际输出目录", manifest=str(manifest_path), output_directory=str(output_dir))

    requested_value = data.get("requested_output_path")
    requested_path = Path(requested_value).expanduser().resolve() if requested_value else None
    if requested_path:
        expected_dir = requested_path.parent if requested_path.suffix.lower() == ".xlsx" else requested_path
        if normalized(expected_dir) != normalized(output_dir):
            add(errors, "OUTPUT_PATH_MISMATCH", "实际输出目录与用户指定路径不一致", requested=str(requested_path), actual=str(output_dir))

    feature_key = str(data.get("feature_key") or "").strip()
    if not feature_key:
        add(errors, "MISSING_FEATURE_KEY", "缺少feature_key，无法判断同功能工作簿是否碎片化")

    workbooks = data.get("workbooks")
    if not isinstance(workbooks, list) or not workbooks:
        add(errors, "MISSING_WORKBOOK_PLAN", "workbooks必须是非空数组")
        workbooks = []

    feature_entries = []
    seen_paths: set[str] = set()
    workbook_paths: dict[str, Path] = {}
    seen_sources: dict[str, str] = {}
    declared_sheets: dict[str, str] = {}

    for index, entry in enumerate(workbooks):
        if not isinstance(entry, dict):
            add(errors, "BAD_WORKBOOK_ENTRY", "工作簿条目必须是对象", index=index)
            continue
        role = entry.get("role")
        if role not in VALID_ROLES:
            add(errors, "BAD_WORKBOOK_ROLE", "role必须为feature/shared/referenced", index=index, role=role)
        if role == "feature":
            feature_entries.append(entry)
            if entry.get("feature_key") != feature_key:
                add(errors, "FEATURE_KEY_MISMATCH", "主功能工作簿的feature_key与清单不一致", index=index)
        elif not str(entry.get("reason") or "").strip():
            add(errors, "MISSING_EXCEPTION_REASON", "公共或引用工作簿必须说明归属依据", index=index)

        action = entry.get("delivery_action")
        if action not in VALID_ACTIONS:
            add(errors, "BAD_DELIVERY_ACTION", "delivery_action必须为created或copied_and_modified", index=index, delivery_action=action)
        if role in {"shared", "referenced"} and action != "copied_and_modified":
            add(errors, "EXISTING_DEPENDENCY_NOT_COPIED", "公共或引用工作簿必须整本复制后在副本中配置测试数据", index=index, role=role)

        raw_path = entry.get("path")
        if not raw_path:
            add(errors, "MISSING_WORKBOOK_PATH", "工作簿条目缺少path", index=index)
            continue
        candidate = Path(raw_path).expanduser()
        workbook_path = candidate.resolve() if candidate.is_absolute() else (output_dir / candidate).resolve()
        path_key = normalized(workbook_path)
        if path_key in seen_paths:
            add(errors, "DUPLICATE_WORKBOOK_ENTRY", "同一工作簿在清单中重复出现", path=str(workbook_path))
        seen_paths.add(path_key)
        workbook_paths[path_key] = workbook_path
        if workbook_path.suffix.lower() != ".xlsx":
            add(errors, "BAD_WORKBOOK_EXTENSION", "配置工作簿必须为xlsx", path=str(workbook_path))
        if normalized(workbook_path.parent) != normalized(output_dir):
            add(errors, "WORKBOOK_OUTSIDE_OUTPUT_DIRECTORY", "最终工作簿不在用户输出目录中", path=str(workbook_path))
        if not workbook_path.is_file():
            add(errors, "WORKBOOK_NOT_FOUND", "清单声明的工作簿不存在", path=str(workbook_path))
            continue

        sheets = entry.get("sheets")
        if not isinstance(sheets, list) or not sheets:
            add(errors, "MISSING_SHEET_LIST", "工作簿条目必须声明本次新增或修改的Sheet", path=str(workbook_path))
            continue
        wb = load_workbook(workbook_path, read_only=True, data_only=False)
        actual_sheets = set(wb.sheetnames)
        for sheet in sheets:
            if sheet not in actual_sheets:
                add(errors, "SHEET_NOT_FOUND", "清单声明的Sheet不在工作簿中", path=str(workbook_path), sheet=sheet)
            previous = declared_sheets.get(sheet)
            if previous and previous != path_key:
                add(errors, "SHEET_SPLIT_ACROSS_WORKBOOKS", "同一Sheet被分配到多个工作簿", sheet=sheet)
            declared_sheets[sheet] = path_key

        source_value = entry.get("source_path")
        if action == "created":
            if source_value:
                add(errors, "CREATED_WORKBOOK_HAS_SOURCE", "created工作簿不应声明source_path；既有工作簿应使用copied_and_modified", path=str(workbook_path))
            wb.close()
            continue

        if not source_value:
            add(errors, "MISSING_SOURCE_PATH", "既有工作簿副本必须声明source_path", path=str(workbook_path))
            wb.close()
            continue

        source_candidate = Path(source_value).expanduser()
        if not source_candidate.is_absolute():
            add(errors, "SOURCE_PATH_NOT_ABSOLUTE", "source_path必须是绝对路径", source_path=str(source_candidate))
        source_path = source_candidate.resolve()
        source_key = normalized(source_path)
        if source_key == path_key:
            add(errors, "SOURCE_OVERWRITTEN", "交付副本不能与源工作簿使用同一路径", path=str(workbook_path))
        previous_target = seen_sources.get(source_key)
        if previous_target and previous_target != path_key:
            add(errors, "SOURCE_COPIED_MULTIPLE_TIMES", "同一源工作簿只能复制成一个交付副本", source_path=str(source_path))
        seen_sources[source_key] = path_key
        if entry.get("copy_scope") != "full_workbook":
            add(errors, "PARTIAL_WORKBOOK_COPY", "既有配置必须整本复制，copy_scope必须为full_workbook", path=str(workbook_path))
        if not source_path.is_file():
            add(errors, "SOURCE_WORKBOOK_NOT_FOUND", "source_path指向的源工作簿不存在", source_path=str(source_path))
            wb.close()
            continue

        source_wb = load_workbook(source_path, read_only=True, data_only=False)
        missing_source_sheets = sorted(set(source_wb.sheetnames) - actual_sheets)
        if missing_source_sheets:
            add(errors, "SOURCE_SHEETS_NOT_PRESERVED", "交付副本没有保留源工作簿的全部Sheet", path=str(workbook_path), missing_sheets=missing_source_sheets)

        test_data = entry.get("test_data")
        if not isinstance(test_data, list) or not test_data:
            add(errors, "MISSING_DEPENDENCY_TEST_DATA", "既有工作簿副本必须配置并声明本系统需要的测试数据", path=str(workbook_path))
            test_data = []

        for test_index, test_row in enumerate(test_data):
            if not isinstance(test_row, dict):
                add(errors, "BAD_TEST_DATA_ENTRY", "test_data条目必须是对象", path=str(workbook_path), test_index=test_index)
                continue
            sheet = test_row.get("sheet")
            row_id = test_row.get("id")
            operation = test_row.get("operation")
            purpose = str(test_row.get("purpose") or "").strip()
            if sheet not in actual_sheets:
                add(errors, "TEST_DATA_SHEET_NOT_FOUND", "测试数据Sheet不在交付副本中", path=str(workbook_path), sheet=sheet)
                continue
            if sheet not in sheets:
                add(errors, "TEST_DATA_SHEET_NOT_DECLARED", "测试数据Sheet必须列入workbooks[].sheets", path=str(workbook_path), sheet=sheet)
            if row_id is None or str(row_id).strip() == "":
                add(errors, "MISSING_TEST_DATA_ID", "测试数据必须声明B列ID", path=str(workbook_path), sheet=sheet, test_index=test_index)
                continue
            if operation not in VALID_TEST_OPERATIONS:
                add(errors, "BAD_TEST_DATA_OPERATION", "测试数据operation必须为added或updated", path=str(workbook_path), sheet=sheet, id=row_id, operation=operation)
                continue
            if not purpose:
                add(errors, "MISSING_TEST_DATA_PURPOSE", "测试数据必须说明本系统测试用途", path=str(workbook_path), sheet=sheet, id=row_id)

            target_rows = rows_for_id(wb[sheet], row_id)
            if len(target_rows) != 1:
                add(errors, "TEST_DATA_ID_NOT_UNIQUE_IN_COPY", "测试ID在交付副本中必须且只能存在一行", path=str(workbook_path), sheet=sheet, id=row_id, matches=len(target_rows))
                continue
            source_rows = rows_for_id(source_wb[sheet], row_id) if sheet in source_wb.sheetnames else []
            if operation == "added" and source_rows:
                add(errors, "TEST_DATA_NOT_ADDED", "added测试ID已存在于源工作簿，不能证明是本次新增", path=str(workbook_path), sheet=sheet, id=row_id)
            if operation == "updated":
                if len(source_rows) != 1:
                    add(errors, "UPDATED_TEST_DATA_NOT_IN_SOURCE", "updated测试ID在源工作簿中必须且只能存在一行", source_path=str(source_path), sheet=sheet, id=row_id, matches=len(source_rows))
                elif target_rows[0] == source_rows[0]:
                    add(errors, "UPDATED_TEST_DATA_UNCHANGED", "updated测试行与源工作簿完全相同，副本未实际配置测试数据", path=str(workbook_path), sheet=sheet, id=row_id)

        source_wb.close()
        wb.close()

    id_allocations = data.get("id_allocations")
    if not isinstance(id_allocations, list) or not id_allocations:
        add(errors, "MISSING_ID_PLAN", "id_allocations必须是非空数组，记录本次新增或修改的每个ID")
        id_allocations = []
    seen_allocations: set[tuple[str, int]] = set()
    for allocation_index, allocation in enumerate(id_allocations):
        if not isinstance(allocation, dict):
            add(errors, "BAD_ID_ALLOCATION", "id_allocations条目必须是对象", index=allocation_index)
            continue
        sheet = allocation.get("sheet")
        row_id = allocation.get("id")
        scope = allocation.get("scope")
        kind = allocation.get("kind")
        source = allocation.get("source")
        status = allocation.get("status")
        rule = str(allocation.get("allocation_rule") or "").strip()
        parent_id = allocation.get("parent_id")
        if not isinstance(sheet, str) or not sheet.strip():
            add(errors, "MISSING_ID_ALLOCATION_SHEET", "ID台账必须声明sheet", index=allocation_index)
            continue
        if not isinstance(row_id, int) or isinstance(row_id, bool) or not (0 < row_id <= INT32_MAX):
            add(errors, "BAD_ID_ALLOCATION_VALUE", "ID台账中的id必须是Int32范围内的正整数", sheet=sheet, id=row_id)
            continue
        if scope not in VALID_ID_SCOPES:
            add(errors, "BAD_ID_ALLOCATION_SCOPE", "ID台账scope必须为sheet/module/parent", sheet=sheet, id=row_id, scope=scope)
        if kind not in VALID_ID_KINDS:
            add(errors, "BAD_ID_ALLOCATION_KIND", "ID台账kind不在允许值内", sheet=sheet, id=row_id, kind=kind)
        if source not in VALID_ID_SOURCES:
            add(errors, "BAD_ID_ALLOCATION_SOURCE", "ID台账source必须为S/T/D/A", sheet=sheet, id=row_id, source=source)
        if status not in VALID_ID_STATUSES:
            add(errors, "BAD_ID_ALLOCATION_STATUS", "ID台账status必须为reused/candidate/confirmed", sheet=sheet, id=row_id, status=status)
        if not rule:
            add(errors, "MISSING_ID_ALLOCATION_RULE", "ID台账必须写明B1或策划构成规则", sheet=sheet, id=row_id)
        if parent_id is not None and (not isinstance(parent_id, int) or isinstance(parent_id, bool) or not (0 < parent_id <= INT32_MAX)):
            add(errors, "BAD_ID_ALLOCATION_PARENT", "parent_id必须为空或Int32范围内的正整数", sheet=sheet, id=row_id, parent_id=parent_id)
        if allocation.get("collision_checked") is not True:
            add(errors, "ID_COLLISION_NOT_CHECKED", "ID台账必须明确collision_checked=true", sheet=sheet, id=row_id)

        workbook_key = declared_sheets.get(sheet)
        if not workbook_key:
            add(errors, "ID_ALLOCATION_SHEET_NOT_DECLARED", "ID台账中的Sheet必须列入工作簿归属计划", sheet=sheet, id=row_id)
            continue
        allocation_key = (sheet, row_id)
        if allocation_key in seen_allocations:
            add(errors, "DUPLICATE_ID_ALLOCATION", "同一Sheet和ID在台账中重复", sheet=sheet, id=row_id)
            continue
        seen_allocations.add(allocation_key)
        workbook_path = workbook_paths[workbook_key]
        wb = load_workbook(workbook_path, read_only=True, data_only=False)
        matches = rows_for_id(wb[sheet], row_id) if sheet in wb.sheetnames else []
        wb.close()
        if len(matches) != 1:
            add(errors, "ID_ALLOCATION_ROW_NOT_FOUND", "ID台账中的ID在交付工作簿中必须且只能存在一行", path=str(workbook_path), sheet=sheet, id=row_id, matches=len(matches))

    if len(feature_entries) != 1:
        add(errors, "FEATURE_WORKBOOK_FRAGMENTATION", "同一个feature_key必须且只能有一个主功能工作簿", feature_key=feature_key, feature_workbook_count=len(feature_entries))

    if requested_path and requested_path.suffix.lower() == ".xlsx" and len(feature_entries) == 1:
        feature_raw = Path(feature_entries[0]["path"]).expanduser()
        feature_path = feature_raw.resolve() if feature_raw.is_absolute() else (output_dir / feature_raw).resolve()
        if normalized(feature_path) != normalized(requested_path):
            add(errors, "FEATURE_FILENAME_MISMATCH", "主功能工作簿没有使用用户指定的完整xlsx路径", requested=str(requested_path), actual=str(feature_path))

    report = {
        "ok": not errors,
        "manifest": str(manifest_path),
        "resolved_output_directory": str(output_dir),
        "feature_key": feature_key,
        "workbook_count": len(workbooks),
        "feature_workbook_count": len(feature_entries),
        "id_allocation_count": len(id_allocations),
        "error_count": len(errors),
        "warning_count": len(warnings),
        "errors": errors,
        "warnings": warnings,
    }
    sys.stdout.write(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())

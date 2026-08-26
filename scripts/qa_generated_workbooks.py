from __future__ import annotations

import argparse
import io
import json
import posixpath
import re
import sys
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("需要 Python 3 和 openpyxl 才能运行此验证器") from exc


ALLOWED_TYPES = {"int", "str", "arr", "bool"}
INT32_MAX = 2_147_483_647
CJK_RE = re.compile(r"[\u3400-\u9fff]")
INTERNAL_HINTS = ("路径", "path", "prefab", "spine", "音频", "特效", "节点", "组件", "sku", "渠道", "程序", "脚本")


def iter_nested_numbers(value: Any) -> Iterable[int]:
    if isinstance(value, bool):
        return
    if isinstance(value, int):
        yield value
    elif isinstance(value, float) and value.is_integer():
        yield int(value)
    elif isinstance(value, list):
        for item in value:
            yield from iter_nested_numbers(item)
    elif isinstance(value, dict):
        for item in value.values():
            yield from iter_nested_numbers(item)


def sheet_end_column(ws) -> int:
    for col in range(2, ws.max_column + 1):
        value = ws.cell(1, col).value
        if isinstance(value, str) and value.strip().lower() == "end":
            return col
    return ws.max_column + 1


def data_rows(ws):
    for row in range(7, ws.max_row + 1):
        marker = ws.cell(row, 1).value
        if isinstance(marker, str) and marker.strip().lower() == "end":
            break
        yield row


XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return ["".join(node.text or "" for node in item.iter(f"{{{XML_NS}}}t")) for item in root]


def _cell_value(cell, shared_strings: list[str]):
    cell_type = cell.attrib.get("t")
    value_node = cell.find(f"{{{XML_NS}}}v")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.iter(f"{{{XML_NS}}}t"))
    if value_node is None or value_node.text is None:
        return None
    raw = value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return raw
    if cell_type in {"str", "e"}:
        return raw
    try:
        number = float(raw)
        return int(number) if number.is_integer() else number
    except ValueError:
        return raw


def _sheet_xml_paths(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall(f"{{{PKG_REL_NS}}}Relationship")
    }
    result = {}
    for sheet in workbook.findall(f".//{{{XML_NS}}}sheet"):
        rel_id = sheet.attrib.get(f"{{{REL_NS}}}id")
        target = targets.get(rel_id, "")
        if target.startswith("/"):
            path = target.lstrip("/")
        else:
            path = posixpath.normpath(posixpath.join("xl", target))
        result[sheet.attrib["name"]] = path
    return result


def collect_ids(paths: list[Path], wanted_sheets: set[str]) -> dict[str, set[int]]:
    ids: dict[str, set[int]] = defaultdict(set)
    if not wanted_sheets:
        return ids
    for path in paths:
        try:
            archive = zipfile.ZipFile(path)
        except Exception:
            continue
        with archive:
            shared_strings = _shared_strings(archive)
            sheet_paths = _sheet_xml_paths(archive)
            for sheet_name in wanted_sheets:
                xml_path = sheet_paths.get(sheet_name)
                if not xml_path or xml_path not in archive.namelist():
                    continue
                stream = io.BytesIO(archive.read(xml_path))
                for _event, row in ET.iterparse(stream, events=("end",)):
                    if row.tag != f"{{{XML_NS}}}row":
                        continue
                    row_number = int(row.attrib.get("r", "0"))
                    if row_number < 7:
                        row.clear()
                        continue
                    marker = None
                    id_value = None
                    for cell in row.findall(f"{{{XML_NS}}}c"):
                        coordinate = cell.attrib.get("r", "")
                        if coordinate.startswith("A"):
                            marker = _cell_value(cell, shared_strings)
                        elif coordinate.startswith("B"):
                            id_value = _cell_value(cell, shared_strings)
                    if isinstance(marker, str) and marker.strip().lower() == "end":
                        row.clear()
                        break
                    if isinstance(id_value, int) and not isinstance(id_value, bool) and id_value > 0:
                        ids[sheet_name].add(id_value)
                    elif isinstance(id_value, float) and id_value.is_integer() and id_value > 0:
                        ids[sheet_name].add(int(id_value))
                    row.clear()
    return ids


def main() -> int:
    skill_root = Path(__file__).resolve().parents[1]
    default_project = Path(__file__).resolve().parents[4]
    default_table = default_project / "策划" / "配置表" / "Table"

    parser = argparse.ArgumentParser(description="验证 Bubble AI 生成配置表")
    parser.add_argument("workbooks", nargs="+", type=Path)
    parser.add_argument("--table-dir", type=Path, default=default_table)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()

    inputs = [path.resolve() for path in args.workbooks]
    missing = [str(path) for path in inputs if not path.is_file()]
    if missing:
        raise SystemExit("输入文件不存在: " + ", ".join(missing))

    relation_path = skill_root / "references" / "relation-dictionary.json"
    relations = json.loads(relation_path.read_text(encoding="utf-8"))
    relation_map: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for relation in relations.get("high_confidence_relations", []):
        relation_map[(relation["source"], relation["field"])].append(relation)

    input_sheet_names: set[str] = set()
    for path in inputs:
        probe = load_workbook(path, read_only=True, data_only=True)
        input_sheet_names.update(name for name in probe.sheetnames if name.startswith("t"))
        probe.close()
    required_targets = {
        relation["target"]
        for (source, _field), field_relations in relation_map.items()
        if source in input_sheet_names
        for relation in field_relations
        if relation.get("target") == "tlanguage_cn" or float(relation.get("coverage") or 0) > 0
    }
    field_dictionary_path = skill_root / "references" / "field-dictionary.json"
    field_dictionary = json.loads(field_dictionary_path.read_text(encoding="utf-8"))
    target_workbook_names = {
        table["source_workbook"]
        for table in field_dictionary.get("tables", [])
        if table.get("sheet") in required_targets
    }
    source_paths = (
        [args.table_dir.resolve() / name for name in sorted(target_workbook_names)]
        if args.table_dir.is_dir()
        else []
    )
    all_ids = collect_ids(source_paths + inputs, required_targets)

    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    checked_sheets = 0
    checked_rows = 0

    def add(bucket, code, path, sheet, cell, message):
        bucket.append({"code": code, "workbook": str(path), "sheet": sheet, "cell": cell, "message": message})

    for path in inputs:
        wb = load_workbook(path, data_only=False)
        cached = load_workbook(path, data_only=True)
        export_sheets = [ws for ws in wb.worksheets if ws.title.startswith("t")]
        if not export_sheets:
            add(errors, "NO_EXPORT_SHEET", path, "", "", "工作簿没有以小写t开头的正式Sheet")

        for ws in export_sheets:
            checked_sheets += 1
            cached_ws = cached[ws.title]
            end_col = sheet_end_column(ws)
            if end_col == ws.max_column + 1:
                add(warnings, "NO_END_COLUMN", path, ws.title, "1", "未找到列END；新表建议显式添加")

            if ws.cell(3, 2).value != "id" or ws.cell(6, 2).value != "int":
                add(errors, "BAD_PRIMARY_HEADER", path, ws.title, "B3/B6", "B3必须为id且B6必须为int")

            field_names: list[str] = []
            fields: dict[int, dict[str, str]] = {}
            for col in range(2, end_col):
                header_values = [ws.cell(row, col).value for row in range(1, 7)]
                if any(value is None or str(value).strip() == "" for value in header_values):
                    add(errors, "INCOMPLETE_HEADER", path, ws.title, ws.cell(1, col).coordinate, "字段第1至6行不得为空")
                    continue
                field_name = str(ws.cell(3, col).value).strip()
                field_type = str(ws.cell(6, col).value).strip()
                client = str(ws.cell(4, col).value).strip()
                server = str(ws.cell(5, col).value).strip()
                comment = str(ws.cell(1, col).value)
                field_names.append(field_name)
                fields[col] = {"name": field_name, "type": field_type, "client": client, "server": server, "comment": comment}
                if field_type not in ALLOWED_TYPES:
                    add(errors, "BAD_TYPE", path, ws.title, ws.cell(6, col).coordinate, f"不支持的类型: {field_type}")
                if client not in {"0", "1"} or server not in {"0", "1"}:
                    add(errors, "BAD_EXPORT_FLAG", path, ws.title, ws.cell(4, col).coordinate, "客户端/服务端开关必须为0或1")

            duplicates = sorted({name for name in field_names if field_names.count(name) > 1})
            if duplicates:
                add(errors, "DUPLICATE_FIELD", path, ws.title, "3", "重复字段名: " + ", ".join(duplicates))

            seen_ids: set[int] = set()
            for row in data_rows(ws):
                raw_id = ws.cell(row, 2).value
                if raw_id in (None, "", 0, "0"):
                    continue
                checked_rows += 1
                if not isinstance(raw_id, int) or isinstance(raw_id, bool) or not (0 < raw_id <= INT32_MAX):
                    add(errors, "BAD_ID", path, ws.title, ws.cell(row, 2).coordinate, "ID必须为Int32范围内的正整数")
                elif raw_id in seen_ids:
                    add(errors, "DUPLICATE_ID", path, ws.title, ws.cell(row, 2).coordinate, f"重复ID: {raw_id}")
                else:
                    seen_ids.add(raw_id)

                exported_nonzero = False
                for col, meta in fields.items():
                    value = ws.cell(row, col).value
                    if value in (None, ""):
                        continue
                    if col != 2 and meta["client"] == "1" and value not in (0, "0", False):
                        exported_nonzero = True
                    coord = ws.cell(row, col).coordinate

                    if isinstance(value, str) and value.startswith("="):
                        if cached_ws.cell(row, col).value is None:
                            add(warnings, "FORMULA_NO_CACHE", path, ws.title, coord, "公式没有可读取缓存值")
                        continue
                    parsed_array = None
                    if meta["type"] == "arr":
                        try:
                            parsed_array = json.loads(str(value))
                            if not isinstance(parsed_array, list):
                                raise ValueError("not array")
                        except Exception:
                            add(errors, "BAD_ARRAY_JSON", path, ws.title, coord, "arr必须为严格JSON数组")
                            continue
                    elif meta["type"] == "bool" and value not in (True, False, 0, 1, "0", "1", "True", "False", "true", "false"):
                        add(errors, "BAD_BOOL", path, ws.title, coord, "bool必须为True/False或0/1")

                    field_relations = relation_map.get((ws.title, meta["name"]), [])
                    if field_relations and value not in (0, "0"):
                        candidates: list[int] = []
                        if parsed_array is not None:
                            candidates = list(iter_nested_numbers(parsed_array))
                        elif isinstance(value, int) and not isinstance(value, bool):
                            candidates = [value]
                        elif isinstance(value, float) and value.is_integer():
                            candidates = [int(value)]
                        for relation in field_relations:
                            target = relation["target"]
                            missing_refs = sorted({candidate for candidate in candidates if candidate > 0 and candidate not in all_ids.get(target, set())})
                            if missing_refs:
                                strict = target == "tlanguage_cn" or float(relation.get("coverage") or 0) >= 0.95
                                bucket = errors if strict else warnings
                                code = "MISSING_REFERENCE" if strict else "UNMATCHED_REFERENCE_CANDIDATE"
                                add(bucket, code, path, ws.title, coord, f"引用{target}未匹配目标主键: {missing_refs[:10]}")

                    if ws.title != "tlanguage_cn" and meta["type"] == "str" and isinstance(value, str) and CJK_RE.search(value):
                        hint = (meta["comment"] + " " + meta["name"]).lower()
                        if not any(token.lower() in hint for token in INTERNAL_HINTS):
                            add(warnings, "POSSIBLE_HARDCODED_TEXT", path, ws.title, coord, "疑似玩家可见中文；应确认是否拆入tlanguage_cn")

                if not exported_nonzero:
                    add(warnings, "ID_ONLY_OR_OMITTED_ROW", path, ws.title, ws.cell(row, 2).coordinate, "除ID外没有客户端可写出的非零/非空字段")

        wb.close()
        cached.close()

    report = {
        "ok": not errors,
        "checked_workbooks": len(inputs),
        "checked_export_sheets": checked_sheets,
        "checked_data_rows": checked_rows,
        "error_count": len(errors),
        "warning_count": len(warnings),
        "errors": errors,
        "warnings": warnings,
    }
    output = json.dumps(report, ensure_ascii=False, indent=2) + "\n"
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(output, encoding="utf-8")
    sys.stdout.write(output)
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
